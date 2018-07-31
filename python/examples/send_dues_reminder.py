#! python3
# send_dues_reminders.py - Sends emails based on their status in spreadsheet.

import openpyxl
import smtplib
import sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.get_highest_column()
latest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}
# Check each member's payment status
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# Log in to email account.
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('example@gmail.com', sys.argv[1])

# Send out reminder emails.
for name, email in unpaid_members.items():
    body = 'Subject: {} dues unpaid.\nDear {},\nRecords show that you have not paid dues for {}. ' \
           'Please make this payment as soon as possible. Thank you!'.format(latest_month, name, latest_month)
    print('Sending email to {}...'.format(email))
    sendmail_status = smtp_obj.sendmail('example@gmail.com', email, body)

    if sendmail_status != {}:
        print('There was a problem sending email to {}: {}'.format(email, sendmail_status))
smtp_obj.quit()
