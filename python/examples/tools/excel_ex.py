import openpyxl
import datetime

wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.active
sheet.title
sheet.title = 'Spam Bacon Eggs Sheet'
wb.get_sheet_names()
wb.save('example2.xlsx')

wb = openpyxl.load_workbook('example.xlsx')
wb.get_sheet_names()
['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
type(sheet)
sheet.title
anotherSheet = wb.active
print(anotherSheet)

sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1']
sheet['A1'].value
datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
c.value
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
sheet['C1'].value
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

sheet.max_row
sheet.max_column

tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

wb.create_sheet(index=2, title='Middle Sheet')
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))

sheet['A1'] = 'Hello world!'

